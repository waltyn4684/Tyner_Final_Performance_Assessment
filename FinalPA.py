# Wallace Tyner

print("waltyn4684")

import os
import openpyxl
import matplotlib.pyplot as plt
from openpyxl.chart import Reference, PieChart

# Change to Final Exam folder
os.chdir(r"C:\SDC205 - FinalExam")


def askUser():

    # Loop through five numbers entered by the user
    # and calculate a running total

    total = 0

    for count in range(5):
        number = int(input("Please enter a number: "))
        total += number

    print("The sum for the 5 numbers entered is:", total)


def askIncome():

    # Loop through five people and collect names
    # and annual income values

    file = open("final.csv", "a")

    for count in range(5):

        name = input("Please enter a name: ")
        income = input("Please enter their income: ")

        file.write(f"\n{name},{income}")

    file.close()


def excelPie():

    # Read names and incomes from CSV

    names = []
    incomes = []

    file = open("final.csv", "r")

    for line in file:

        row = line.strip().split(",")

        # Skip blank rows or bad rows
        if len(row) < 2:
            continue

        names.append(row[0])
        incomes.append(int(row[1]))

    file.close()

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write data to worksheet
    for i in range(len(names)):
        ws.cell(row=i + 1, column=1).value = names[i]
        ws.cell(row=i + 1, column=2).value = incomes[i]

    # Create chart labels
    labels = Reference(
        ws,
        min_col=1,
        min_row=1,
        max_row=len(names)
    )

    # Create chart data
    data = Reference(
        ws,
        min_col=2,
        min_row=1,
        max_row=len(names)
    )

    # Create pie chart
    pie = PieChart()

    # Set chart title
    pie.title = "waltyn4684"

    # Add data and labels
    pie.add_data(data)
    pie.set_categories(labels)

    # Place chart
    ws.add_chart(pie, "D1")

    # Save workbook
    wb.save("final.xlsx")


def verticalBar():

    # Read CSV data for graph

    names = []
    incomes = []

    file = open("final.csv", "r")

    for line in file:

        row = line.strip().split(",")

        # Skip blank rows or bad rows
        if len(row) < 2:
            continue

        names.append(row[0])
        incomes.append(int(row[1]))

    file.close()

    # Create bar graph
    plt.bar(
        names,
        incomes,
        color="green",
        label="Income"
    )

    plt.xlabel("Name")
    plt.ylabel("Income")
    plt.title("waltyn4684")
    plt.legend()

    plt.show()


# Run all functions

askUser()

askIncome()

excelPie()

verticalBar()
